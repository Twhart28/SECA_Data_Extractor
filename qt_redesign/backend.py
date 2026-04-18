from __future__ import annotations

import math
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PIL import Image, ImageOps, ImageStat

MEASUREMENT_BASE_WIDTH = 9917
MEASUREMENT_BASE_HEIGHT = 14034
OCR_RENDER_RESOLUTION = 300
MEASUREMENT_CROP_BOXES = [
    (("Fat Mass (kg)", "Fat Mass (%)"), (7045, 3697, 8318, 4029)),
    (("Fat Mass Index (kg/m^2)",), (7045, 4029, 8318, 4543)),
    (("Fat-Free Mass (kg)", "Fat-Free Mass (%)"), (7045, 4543, 8318, 4857)),
    (("Fat-Free Mass Index (kg/m^2)",), (7045, 4857, 8318, 5365)),
    (("Skeletal Muscle Mass (kg)",), (7045, 5365, 8318, 5703)),
    (("Right Arm (kg)",), (7045, 5703, 8318, 6205)),
    (("Left Arm (kg)",), (7045, 6205, 8318, 6547)),
    (("Right Leg (kg)",), (7045, 6547, 8318, 7055)),
    (("Left Leg (kg)",), (7045, 7055, 8318, 7396)),
    (("Torso (kg)",), (7045, 7396, 8318, 7882)),
    (("Visceral Adipose Tissue",), (7045, 7882, 8318, 8218)),
    (("SECA BMI (kg/m^2)",), (7045, 8218, 8318, 8750)),
    (("Height (m)",), (7045, 8750, 8318, 9087)),
    (("Weight (kg)",), (7045, 9087, 8318, 9564)),
    (("Total Body Water (L)", "Total Body Water (%)"), (7045, 9564, 8318, 9908)),
    (("Extracellular Water (L)", "Extracellular Water (%)"), (7045, 9908, 8318, 10456)),
    (("ECW/TBW (%)",), (7045, 10456, 8318, 10788)),
    (("Resting Energy Expenditure (kcal/day)",), (7045, 10788, 8318, 11296)),
    (("Energy Consumption (kcal/day)",), (7045, 11296, 8318, 11636)),
    (("Phase Angle (deg)", "Phase Angle Percentile"), (7045, 11636, 8318, 12150)),
    (("Resistance (Ohm)",), (7045, 12150, 8318, 12486)),
    (("Reactance (Ohm)",), (7045, 12486, 8318, 12990)),
    (("Physical Activity Level",), (7045, 12990, 8318, 13325)),
]

DEFAULT_TESSERACT_DIR = Path(
    os.environ.get("SECA_TESSERACT_DIR", r"C:\Program Files\Tesseract-OCR")
)
_OCR_CONFIG: Optional[str] = None
_OCR_RUNTIME_VALIDATED = False

NUMBER_PATTERN = re.compile(r"-?\d+(?:[.,]\d+)?")
PATIENT_FIELDS = {
    "Age": re.compile(r"\bAge[:\s]+(\d+)", re.IGNORECASE),
}
PATIENT_METADATA_FIELDS = [
    "Scanned ID",
    "Sex",
    "Age",
    "Collection Date",
    "Collection Time",
]
MEASUREMENT_FIELD_NAMES: List[str] = [
    "Fat Mass (kg)",
    "Fat Mass (%)",
    "Fat Mass Index (kg/m^2)",
    "Fat-Free Mass (kg)",
    "Fat-Free Mass (%)",
    "Fat-Free Mass Index (kg/m^2)",
    "Skeletal Muscle Mass (kg)",
    "Right Arm (kg)",
    "Left Arm (kg)",
    "Right Leg (kg)",
    "Left Leg (kg)",
    "Torso (kg)",
    "Visceral Adipose Tissue",
    "SECA BMI (kg/m^2)",
    "Height (m)",
    "Weight (kg)",
    "Total Body Water (L)",
    "Total Body Water (%)",
    "Extracellular Water (L)",
    "Extracellular Water (%)",
    "ECW/TBW (%)",
    "Resting Energy Expenditure (kcal/day)",
    "Energy Consumption (kcal/day)",
    "Phase Angle (deg)",
    "Phase Angle Percentile",
    "Resistance (Ohm)",
    "Reactance (Ohm)",
    "Physical Activity Level",
]
CALCULATED_FIELD_NAMES: List[str] = [
    "Body Mass Index (kg/m^2)",
]
REVIEWABLE_FIELDS = set(MEASUREMENT_FIELD_NAMES + CALCULATED_FIELD_NAMES)
DEBUG_SAVE_OCR_TXT = 0


def _runtime_search_roots() -> List[Path]:
    roots: List[Path] = []

    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        roots.extend([exe_dir, exe_dir / "_internal"])
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            roots.append(Path(meipass))

    roots.append(Path(__file__).resolve().parent)
    roots.append(DEFAULT_TESSERACT_DIR)

    deduped: List[Path] = []
    seen: set[str] = set()
    for root in roots:
        key = str(root).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(root)
    return deduped


def resolve_tesseract_runtime() -> tuple[Path, Path]:
    env_cmd = os.environ.get("SECA_TESSERACT_CMD")
    if env_cmd:
        cmd_path = Path(env_cmd).expanduser()
        tessdata_dir = cmd_path.parent / "tessdata"
        if cmd_path.exists() and tessdata_dir.exists():
            return cmd_path, tessdata_dir

    for root in _runtime_search_roots():
        cmd_path = root / "tesseract" / "tesseract.exe"
        tessdata_dir = root / "tesseract" / "tessdata"
        if cmd_path.exists() and tessdata_dir.exists():
            return cmd_path, tessdata_dir

        cmd_path = root / "tesseract.exe"
        tessdata_dir = root / "tessdata"
        if cmd_path.exists() and tessdata_dir.exists():
            return cmd_path, tessdata_dir

    which_path = shutil.which("tesseract")
    if which_path:
        cmd_path = Path(which_path)
        tessdata_dir = cmd_path.parent / "tessdata"
        if tessdata_dir.exists():
            return cmd_path, tessdata_dir

    raise FileNotFoundError(
        "Tesseract OCR runtime was not found. Use the packaged build, or install "
        "Tesseract locally, or set SECA_TESSERACT_CMD / SECA_TESSERACT_DIR."
    )


def ensure_tesseract_runtime() -> str:
    global _OCR_CONFIG, _OCR_RUNTIME_VALIDATED

    if _OCR_CONFIG is None:
        cmd_path, tessdata_dir = resolve_tesseract_runtime()
        pytesseract.pytesseract.tesseract_cmd = str(cmd_path)
        os.environ["TESSDATA_PREFIX"] = str(tessdata_dir)
        _OCR_CONFIG = ""

    if not _OCR_RUNTIME_VALIDATED:
        cmd_path = Path(pytesseract.pytesseract.tesseract_cmd)
        process_kwargs = {
            "stdout": subprocess.PIPE,
            "stderr": subprocess.STDOUT,
            "stdin": subprocess.DEVNULL,
            "env": os.environ.copy(),
            "text": True,
        }
        if hasattr(subprocess, "STARTUPINFO"):
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
            process_kwargs["startupinfo"] = startupinfo
        try:
            subprocess.run(
                [str(cmd_path), "--version"],
                check=True,
                **process_kwargs,
            )
        except (OSError, subprocess.SubprocessError) as exc:
            raise pytesseract.TesseractNotFoundError() from exc
        _OCR_RUNTIME_VALIDATED = True

    return _OCR_CONFIG


def image_content_score(image: Image.Image) -> float:
    gray_image = image.convert("L")
    inverted = ImageOps.invert(gray_image)
    bbox = inverted.getbbox()
    if bbox is None:
        return 0.0

    width = max(1, gray_image.width)
    height = max(1, gray_image.height)
    bbox_area = max(1, bbox[2] - bbox[0]) * max(1, bbox[3] - bbox[1])
    area_ratio = bbox_area / float(width * height)
    stats = ImageStat.Stat(gray_image)
    contrast = float(stats.var[0]) / 1000.0
    darkness = max(0.0, 255.0 - float(stats.mean[0])) / 255.0
    return area_ratio + contrast + darkness


def output_field_order() -> List[str]:
    order: List[str] = []
    for name in MEASUREMENT_FIELD_NAMES:
        order.append(name)
        if name == "SECA BMI (kg/m^2)":
            order.extend(CALCULATED_FIELD_NAMES)
    return [
        "Source File",
        "Patient ID",
        *PATIENT_METADATA_FIELDS,
        "Data Quality",
        "Data Quality Fails",
        *order,
    ]


OUTPUT_FIELD_ORDER = output_field_order()


def scale_box_to_image(box, image_size):
    img_w, img_h = image_size
    sx = img_w / MEASUREMENT_BASE_WIDTH
    sy = img_h / MEASUREMENT_BASE_HEIGHT
    x0, y0, x1, y1 = box
    return (
        int(x0 * sx),
        int(y0 * sy),
        int(x1 * sx),
        int(y1 * sy),
    )


def normalize_number(token: str) -> float:
    return float(token.replace(",", "."))


def collapse_whitespace(text: str) -> str:
    return " ".join(text.split())


def normalized_text_token(value: Optional[str]) -> str:
    return (value or "").strip().casefold()


def normalize_row_precision(row: Dict[str, object]) -> None:
    bmi_value = row.get("Body Mass Index (kg/m^2)")
    if isinstance(bmi_value, (int, float)):
        row["Body Mass Index (kg/m^2)"] = round(float(bmi_value), 2)


def extract_patient_id_from_filename(pdf_path: Path) -> Optional[str]:
    stem = pdf_path.stem.strip()
    if not stem:
        return None

    first_break_index: Optional[int] = None
    for index, char in enumerate(stem):
        if char in (" ", "_"):
            first_break_index = index
            break

    if first_break_index is None:
        return stem

    next_char = stem[first_break_index + 1:first_break_index + 2]
    if next_char.upper() == "T":
        for index in range(first_break_index + 1, len(stem)):
            if stem[index] in (" ", "_"):
                return stem[:index].strip() or None
        return stem or None

    return stem[:first_break_index].strip() or None


def resolve_scanned_id(header_text: str) -> Optional[str]:
    id_name_match = re.search(r"ID\s*[:\-]?\s*(.*?)\s+Name", header_text, re.IGNORECASE)
    legacy_scanned_id = id_name_match.group(1).strip() if id_name_match else None

    name_match = re.search(r"Name:\s*(\S+)", header_text, re.IGNORECASE)
    name_token = name_match.group(1).strip() if name_match else None
    if name_token and not re.search(r"[A-Za-z]", name_token):
        name_token = None

    if legacy_scanned_id and "seca_" in legacy_scanned_id.casefold():
        return name_token or None

    if legacy_scanned_id and name_token:
        if normalized_text_token(legacy_scanned_id) == normalized_text_token(name_token):
            return name_token
        return name_token

    if name_token:
        return name_token

    return legacy_scanned_id or None


def extract_measurements_from_page_image(
    pil_image: Image.Image,
) -> Tuple[Dict[str, Optional[float]], List[str], Dict[str, object], Dict[str, float]]:
    measurements: Dict[str, Optional[float]] = {}
    debug_lines: List[str] = []
    field_images: Dict[str, object] = {}
    field_image_scores: Dict[str, float] = {}
    ocr_config = ensure_tesseract_runtime()

    for fields, base_box in MEASUREMENT_CROP_BOXES:
        try:
            crop_box = scale_box_to_image(base_box, pil_image.size)
            cropped = pil_image.crop(crop_box)
            ocr_text = pytesseract.image_to_string(cropped, config=ocr_config)
        except Exception:
            cropped = None
            ocr_text = ""

        cleaned_text = collapse_whitespace(ocr_text)
        if cleaned_text:
            debug_lines.append(f"{' | '.join(fields)} => {cleaned_text}")

        numbers = [normalize_number(match) for match in NUMBER_PATTERN.findall(ocr_text)]
        for index, field in enumerate(fields):
            measurements[field] = numbers[index] if index < len(numbers) else None
            if field not in field_images and cropped is not None:
                field_images[field] = cropped.copy()
            if cropped is not None:
                field_image_scores[field] = image_content_score(cropped) + (
                    10.0 if measurements[field] is not None else 0.0
                )

    return measurements, debug_lines, field_images, field_image_scores


def extract_measurements_from_pdf(
    pdf_path: Path,
) -> Tuple[Dict[str, Optional[float]], str, Dict[str, object]]:
    measurements: Dict[str, Optional[float]] = {name: None for name in MEASUREMENT_FIELD_NAMES}
    debug_parts: List[str] = []
    field_images: Dict[str, object] = {}
    field_image_scores: Dict[str, float] = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            pil_image = page.to_image(resolution=OCR_RENDER_RESOLUTION).original
            page_measurements, page_debug_lines, page_field_images, page_field_scores = extract_measurements_from_page_image(
                pil_image
            )
            for field, value in page_measurements.items():
                if value is not None:
                    measurements[field] = value
            for field, image in page_field_images.items():
                candidate_score = page_field_scores.get(field, 0.0)
                current_score = field_image_scores.get(field, -1.0)
                if candidate_score > current_score:
                    field_images[field] = image
                    field_image_scores[field] = candidate_score
            if page_debug_lines:
                debug_parts.append(f"Page {page_index}:\n" + "\n".join(page_debug_lines))

    if not debug_parts and not any(value is not None for value in measurements.values()):
        raise RuntimeError(
            "OCR did not detect any measurement values in this PDF. This usually means the "
            "report layout changed, the wrong page was rendered, or the OCR runtime is not working."
        )

    return measurements, "\n\n".join(debug_parts), field_images


def extract_text_layer(pdf_path: Path) -> str:
    parts: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def normalize_collection_date(raw_date: Optional[str]) -> Optional[str]:
    if not raw_date:
        return None

    cleaned = raw_date.strip()
    parts = re.split(r"[./-]", cleaned)
    if len(parts) == 3:
        try:
            month, day, year = (int(part) for part in parts)
            if year < 100:
                year += 2000 if year < 50 else 1900
            return datetime(year, month, day).strftime("%m/%d/%y")
        except ValueError:
            pass
    return cleaned or None


def normalize_collection_time(raw_time: Optional[str]) -> Optional[str]:
    if not raw_time:
        return None

    match = re.match(r"\s*(\d{1,2}):(\d{2})(?:\s*(AM|PM))?\s*$", raw_time, re.IGNORECASE)
    if not match:
        return raw_time.strip() or None

    hour_text, minute_text, meridiem = match.groups()
    try:
        hour = int(hour_text)
        minute = int(minute_text)
        if not 0 <= minute < 60:
            raise ValueError("Minute out of range")
        if meridiem:
            meridiem = meridiem.upper()
            if hour == 12:
                hour = 0
            if meridiem == "PM":
                hour += 12
        if not 0 <= hour < 24:
            raise ValueError("Hour out of range")
        return f"{hour:02d}:{minute:02d}"
    except ValueError:
        return raw_time.strip() or None


def parse_patient_metadata(text: str) -> Dict[str, Optional[str]]:
    metadata: Dict[str, Optional[str]] = {
        "Scanned ID": resolve_scanned_id(text),
        "Sex": None,
        "Age": None,
        "Collection Date": None,
        "Collection Time": None,
    }

    for field, pattern in PATIENT_FIELDS.items():
        match = pattern.search(text)
        if match:
            metadata[field] = match.group(1).strip()

    sex_match = re.search(r"\b(Male|Female)\b", text, re.IGNORECASE)
    if sex_match:
        metadata["Sex"] = sex_match.group(1).title()

    age_fallback = re.search(r"\b(\d{1,3})\s+(Male|Female)\b", text, re.IGNORECASE)
    if metadata["Age"] is None and age_fallback:
        metadata["Age"] = age_fallback.group(1)

    date_match = re.search(r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text)
    if date_match:
        metadata["Collection Date"] = normalize_collection_date(date_match.group(1))

    time_match = re.search(r"(\d{1,2}:\d{2}\s?(?:AM|PM)?)", text, re.IGNORECASE)
    if time_match:
        metadata["Collection Time"] = normalize_collection_time(time_match.group(1))

    return metadata


def evaluate_data_quality(values: Dict[str, Optional[float]]) -> Dict[str, Optional[str]]:
    def numbers_present(fields: List[str]) -> bool:
        return all(values.get(field) is not None for field in fields)

    def almost_equal(calculated: float, expected: float, tolerance: float) -> bool:
        return abs(calculated - expected) <= tolerance

    def add_decimal_between_first_two_digits(value: Optional[float]) -> Optional[float]:
        if value is None:
            return None
        if not math.isclose(value, round(value), abs_tol=1e-6):
            return None
        sign = -1 if value < 0 else 1
        integer_part = str(int(abs(round(value))))
        if len(integer_part) < 2:
            return None
        return sign * float(f"{integer_part[0]}.{integer_part[1:]}")

    failures: List[str] = []

    if numbers_present(["Fat Mass (kg)", "Fat-Free Mass (kg)", "Weight (kg)"]):
        if not almost_equal((values["Fat Mass (kg)"] or 0) + (values["Fat-Free Mass (kg)"] or 0), values["Weight (kg)"] or 0, 0.01):
            failures.append("1")
    else:
        failures.append("1")

    if numbers_present(["Fat Mass (%)", "Fat-Free Mass (%)"]):
        if not almost_equal((values["Fat Mass (%)"] or 0) + (values["Fat-Free Mass (%)"] or 0), 100, 0.01):
            failures.append("2")
    else:
        failures.append("2")

    if numbers_present(["Fat Mass Index (kg/m^2)", "Fat-Free Mass Index (kg/m^2)", "SECA BMI (kg/m^2)"]):
        if not almost_equal((values["Fat Mass Index (kg/m^2)"] or 0) + (values["Fat-Free Mass Index (kg/m^2)"] or 0), values["SECA BMI (kg/m^2)"] or 0, 0.02):
            failures.append("3")
    else:
        failures.append("3")

    if numbers_present(["Right Arm (kg)", "Left Arm (kg)", "Right Leg (kg)", "Left Leg (kg)", "Torso (kg)", "Skeletal Muscle Mass (kg)"]):
        sum_limbs = sum(values.get(field, 0) or 0 for field in ["Right Arm (kg)", "Left Arm (kg)", "Right Leg (kg)", "Left Leg (kg)", "Torso (kg)"])
        if not almost_equal(sum_limbs, values["Skeletal Muscle Mass (kg)"] or 0, 0.03):
            failures.append("4")
    else:
        failures.append("4")

    if numbers_present(["Weight (kg)", "Height (m)", "SECA BMI (kg/m^2)"]):
        if values["Height (m)"] in (0, None):
            failures.append("5")
        elif not almost_equal((values["Weight (kg)"] or 0) / ((values["Height (m)"] or 1) ** 2), values["SECA BMI (kg/m^2)"] or 0, 0.3):
            failures.append("5")
    else:
        failures.append("5")

    if numbers_present(["Extracellular Water (L)", "Total Body Water (L)", "ECW/TBW (%)"]):
        if values["Total Body Water (L)"] in (0, None):
            failures.append("6")
        elif not almost_equal(((values["Extracellular Water (L)"] or 0) / (values["Total Body Water (L)"] or 1)) * 100, values["ECW/TBW (%)"] or 0, 0.025):
            failures.append("6")
    else:
        failures.append("6")

    if numbers_present(["Extracellular Water (%)", "Total Body Water (%)", "ECW/TBW (%)"]):
        if values["Total Body Water (%)"] in (0, None):
            failures.append("7")
        elif not almost_equal(((values["Extracellular Water (%)"] or 0) / (values["Total Body Water (%)"] or 1)) * 100, values["ECW/TBW (%)"] or 0, 0.02):
            failures.append("7")
    else:
        failures.append("7")

    if numbers_present(["Resting Energy Expenditure (kcal/day)", "Physical Activity Level", "Energy Consumption (kcal/day)"]):
        if not almost_equal((values["Resting Energy Expenditure (kcal/day)"] or 0) * (values["Physical Activity Level"] or 0), values["Energy Consumption (kcal/day)"] or 0, 0.02):
            failures.append("8")
    else:
        failures.append("8")

    if numbers_present(["Reactance (Ohm)", "Resistance (Ohm)", "Phase Angle (deg)"]):
        phase_angle = values["Phase Angle (deg)"]
        original_phase_angle = phase_angle
        if values["Resistance (Ohm)"] in (0, None):
            failures.append("9")
        else:
            calculated = math.atan((values["Reactance (Ohm)"] or 0) / (values["Resistance (Ohm)"] or 1)) * 180 / math.pi
            if not almost_equal(calculated, phase_angle or 0, 0.1):
                adjusted_phase_angle = add_decimal_between_first_two_digits(phase_angle)
                if adjusted_phase_angle is not None and almost_equal(calculated, adjusted_phase_angle, 0.1):
                    values["Phase Angle (deg)"] = adjusted_phase_angle
                else:
                    values["Phase Angle (deg)"] = original_phase_angle
                    failures.append("9")
    else:
        failures.append("9")

    percentile = values.get("Phase Angle Percentile")
    if percentile is None or percentile < 0 or percentile > 100:
        failures.append("10")

    return {
        "Data Quality": "Pass" if not failures else "Fail",
        "Data Quality Fails": ",".join(failures) if failures else "",
    }


QC_FIELD_MAP = {
    "1": ["Fat Mass (kg)", "Fat-Free Mass (kg)", "Weight (kg)"],
    "2": ["Fat Mass (%)", "Fat-Free Mass (%)"],
    "3": ["Fat Mass Index (kg/m^2)", "Fat-Free Mass Index (kg/m^2)", "SECA BMI (kg/m^2)"],
    "4": ["Right Arm (kg)", "Left Arm (kg)", "Right Leg (kg)", "Left Leg (kg)", "Torso (kg)", "Skeletal Muscle Mass (kg)"],
    "5": ["Weight (kg)", "Height (m)", "SECA BMI (kg/m^2)"],
    "6": ["Extracellular Water (L)", "Total Body Water (L)", "ECW/TBW (%)"],
    "7": ["Extracellular Water (%)", "Total Body Water (%)", "ECW/TBW (%)"],
    "8": ["Resting Energy Expenditure (kcal/day)", "Physical Activity Level", "Energy Consumption (kcal/day)"],
    "9": ["Reactance (Ohm)", "Resistance (Ohm)", "Phase Angle (deg)"],
    "10": ["Phase Angle Percentile"],
}


def recompute_calculated_fields(row: Dict[str, Optional[float]]) -> None:
    weight = row.get("Weight (kg)")
    height = row.get("Height (m)")
    if height not in (None, 0):
        row["Body Mass Index (kg/m^2)"] = round(((weight or 0) / ((height or 1) ** 2)), 2) if weight is not None else None
    else:
        row["Body Mass Index (kg/m^2)"] = None


def refresh_data_quality(row: Dict[str, Optional[float]]) -> None:
    recompute_calculated_fields(row)
    row.update(evaluate_data_quality(row))
    normalize_row_precision(row)


def parse_user_value(field_name: str, value: str) -> Optional[object]:
    if field_name in MEASUREMENT_FIELD_NAMES + CALCULATED_FIELD_NAMES:
        cleaned = value.strip()
        if not cleaned:
            return None
        match = NUMBER_PATTERN.search(cleaned)
        return normalize_number(match.group(0)) if match else None
    return value if value.strip() else None


def extract_pdf_data(
    pdf_path: Path, save_ocr_txt: bool = bool(DEBUG_SAVE_OCR_TXT)
) -> Tuple[Dict[str, object], Dict[str, object]]:
    row: Dict[str, object] = {field: None for field in OUTPUT_FIELD_ORDER}
    row["Source File"] = pdf_path.name
    row["Patient ID"] = extract_patient_id_from_filename(pdf_path)

    text_layer = extract_text_layer(pdf_path)
    keyword_text = text_layer.lower()
    if "patient data" not in keyword_text or "single measurement" not in keyword_text:
        row.update(
            {
                "Data Quality": "Fail",
                "Data Quality Fails": "Not recognized as a SECA data export",
            }
        )
        return row, {}

    normalized_header_text = collapse_whitespace(text_layer)
    measurements, ocr_debug_text, field_images = extract_measurements_from_pdf(pdf_path)

    if save_ocr_txt:
        pdf_path.with_suffix(".ocr.txt").write_text(ocr_debug_text, encoding="utf-8")

    row.update(parse_patient_metadata(normalized_header_text))
    row.update(measurements)
    recompute_calculated_fields(row)
    row.update(evaluate_data_quality(row))
    normalize_row_precision(row)
    return row, field_images


def center_text_cells(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    if "All Data" not in workbook.sheetnames:
        return

    worksheet = workbook["All Data"]
    centered = Alignment(horizontal="center", vertical="center")
    skip_columns = {1, 2, 3, 9}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.col_idx in skip_columns:
                continue
            if 10 <= cell.col_idx <= 38:
                cell.alignment = centered
                continue
            if isinstance(cell.value, str) and cell.value != "":
                cell.alignment = centered
    workbook.save(output_path)


def export_entries(entries: List[Dict[str, object]], output_path: Path) -> None:
    rows = [entry["row"] for entry in entries]
    for row in rows:
        normalize_row_precision(row)
    df = pd.DataFrame(rows, columns=OUTPUT_FIELD_ORDER)
    df["__sort_index"] = range(len(df))
    df["__unrecognized"] = df["Data Quality Fails"] == "Not recognized as a SECA data export"
    df = df.sort_values(by=["__unrecognized", "__sort_index"], kind="stable")
    df = df.drop(columns=["__sort_index", "__unrecognized"])
    df.to_excel(output_path, index=False, sheet_name="All Data")
    center_text_cells(output_path)
